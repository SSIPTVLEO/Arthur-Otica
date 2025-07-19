from flask import Blueprint, request, jsonify
from src.models.armacao_lente import db, ArmacaoLente
import openpyxl
import os

armacao_lente_bp = Blueprint('armacao_lente', __name__)

def save_armacoes_to_excel():
    """Salva todos os dados de armações e lentes para o arquivo Excel"""
    try:
        excel_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'SISTEMA-ARTHURÓTICA.xlsx')
        
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook['Armação e lente']
        
        # Limpa dados existentes (mantém apenas o cabeçalho)
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.value = None
        
        # Busca todos os registros do banco
        registros = ArmacaoLente.query.all()
        
        # Adiciona os dados ao Excel
        for i, registro in enumerate(registros, start=2):
            sheet.cell(row=i, column=1, value=registro.id)
            sheet.cell(row=i, column=2, value=registro.id_os)
            sheet.cell(row=i, column=3, value=registro.ponte)
            sheet.cell(row=i, column=4, value=registro.horizontal)
            sheet.cell(row=i, column=5, value=registro.vertical)
            sheet.cell(row=i, column=6, value=registro.diagonal_maior)
            sheet.cell(row=i, column=7, value=registro.marca_armacao)
            sheet.cell(row=i, column=8, value=registro.referencia_armacao)
            sheet.cell(row=i, column=9, value=registro.material_armacao)
            sheet.cell(row=i, column=10, value=registro.lente_comprada)
            sheet.cell(row=i, column=11, value=registro.tratamento)
            sheet.cell(row=i, column=12, value=registro.coloracao)
        
        workbook.save(excel_path)
        return True
    except Exception as e:
        print(f"Erro ao salvar armações no Excel: {e}")
        return False

@armacao_lente_bp.route('/armacoes', methods=['GET'])
def get_armacoes():
    """Retorna todas as armações e lentes"""
    armacoes = ArmacaoLente.query.all()
    return jsonify([armacao.to_dict() for armacao in armacoes])

@armacao_lente_bp.route('/armacoes', methods=['POST'])
def create_armacao():
    """Cria uma nova armação e lente"""
    try:
        data = request.get_json()
        
        # Validação básica
        if 'id_os' not in data or not data['id_os']:
            return jsonify({'error': 'Campo id_os é obrigatório'}), 400
        
        # Cria nova armação
        nova_armacao = ArmacaoLente(
            id_os=int(data['id_os']),
            ponte=float(data.get('ponte')) if data.get('ponte') else None,
            horizontal=float(data.get('horizontal')) if data.get('horizontal') else None,
            vertical=float(data.get('vertical')) if data.get('vertical') else None,
            diagonal_maior=float(data.get('diagonal_maior')) if data.get('diagonal_maior') else None,
            marca_armacao=data.get('marca_armacao'),
            referencia_armacao=data.get('referencia_armacao'),
            material_armacao=data.get('material_armacao'),
            lente_comprada=data.get('lente_comprada'),
            tratamento=data.get('tratamento'),
            coloracao=data.get('coloracao')
        )
        
        db.session.add(nova_armacao)
        db.session.commit()
        
        # Salva no Excel
        if save_armacoes_to_excel():
            return jsonify({'message': 'Armação criada e salva no Excel com sucesso', 'armacao': nova_armacao.to_dict()}), 201
        else:
            return jsonify({'message': 'Armação criada no banco, mas erro ao salvar no Excel', 'armacao': nova_armacao.to_dict()}), 201
            
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@armacao_lente_bp.route('/armacoes/<int:armacao_id>', methods=['GET'])
def get_armacao(armacao_id):
    """Retorna uma armação específica"""
    armacao = ArmacaoLente.query.get_or_404(armacao_id)
    return jsonify(armacao.to_dict())

@armacao_lente_bp.route('/armacoes/<int:armacao_id>', methods=['PUT'])
def update_armacao(armacao_id):
    """Atualiza uma armação e lente"""
    try:
        armacao = ArmacaoLente.query.get_or_404(armacao_id)
        data = request.get_json()
        
        # Atualiza campos se fornecidos
        numeric_fields = ['ponte', 'horizontal', 'vertical', 'diagonal_maior']
        text_fields = ['marca_armacao', 'referencia_armacao', 'material_armacao', 'lente_comprada', 'tratamento', 'coloracao']
        
        if 'id_os' in data:
            armacao.id_os = int(data['id_os'])
        
        for field in numeric_fields:
            if field in data:
                setattr(armacao, field, float(data[field]) if data[field] else None)
        
        for field in text_fields:
            if field in data:
                setattr(armacao, field, data[field])
        
        db.session.commit()
        
        # Salva no Excel
        if save_armacoes_to_excel():
            return jsonify({'message': 'Armação atualizada e salva no Excel com sucesso', 'armacao': armacao.to_dict()})
        else:
            return jsonify({'message': 'Armação atualizada no banco, mas erro ao salvar no Excel', 'armacao': armacao.to_dict()})
            
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@armacao_lente_bp.route('/armacoes/<int:armacao_id>', methods=['DELETE'])
def delete_armacao(armacao_id):
    """Deleta uma armação e lente"""
    try:
        armacao = ArmacaoLente.query.get_or_404(armacao_id)
        db.session.delete(armacao)
        db.session.commit()
        
        # Salva no Excel
        if save_armacoes_to_excel():
            return jsonify({'message': 'Armação deletada e Excel atualizado com sucesso'})
        else:
            return jsonify({'message': 'Armação deletada do banco, mas erro ao atualizar Excel'})
            
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

