from flask import Blueprint, request, jsonify
from src.models.arthurotica import db, Arthurotica
import openpyxl
import os

arthurotica_bp = Blueprint('arthurotica', __name__)

def save_to_excel():
    """Salva todos os dados do banco para o arquivo Excel"""
    try:
        excel_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'SISTEMA-ARTHURÓTICA.xlsx')
        
        # Carrega o workbook existente
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active
        
        # Limpa dados existentes (mantém apenas o cabeçalho)
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.value = None
        
        # Busca todos os registros do banco
        registros = Arthurotica.query.all()
        
        # Adiciona os dados ao Excel
        for i, registro in enumerate(registros, start=2):
            sheet.cell(row=i, column=1, value=registro.id)
            sheet.cell(row=i, column=2, value=registro.id_os)
            sheet.cell(row=i, column=3, value=registro.valor_lente)
            sheet.cell(row=i, column=4, value=registro.valor_armacao)
            sheet.cell(row=i, column=5, value=registro.valor_total)
            sheet.cell(row=i, column=6, value=registro.forma_pagamento)
            sheet.cell(row=i, column=7, value=registro.entrada)
            sheet.cell(row=i, column=8, value=registro.parcelas)
            sheet.cell(row=i, column=9, value=registro.valor_parcelas)
            sheet.cell(row=i, column=10, value=registro.status)
        
        # Salva o arquivo
        workbook.save(excel_path)
        return True
    except Exception as e:
        print(f"Erro ao salvar no Excel: {e}")
        return False

@arthurotica_bp.route('/registros', methods=['GET'])
def get_registros():
    """Retorna todos os registros"""
    registros = Arthurotica.query.all()
    return jsonify([registro.to_dict() for registro in registros])

@arthurotica_bp.route('/registros', methods=['POST'])
def create_registro():
    """Cria um novo registro"""
    try:
        data = request.get_json()
        
        # Validação básica
        required_fields = ['id_os', 'valor_lente', 'valor_armacao', 'forma_pagamento', 'status']
        for field in required_fields:
            if field not in data:
                return jsonify({'error': f'Campo {field} é obrigatório'}), 400
        
        # Calcula valor total
        valor_total = float(data['valor_lente']) + float(data['valor_armacao'])
        
        # Cria novo registro
        novo_registro = Arthurotica(
            id_os=data['id_os'],
            valor_lente=float(data['valor_lente']),
            valor_armacao=float(data['valor_armacao']),
            valor_total=valor_total,
            forma_pagamento=data['forma_pagamento'],
            entrada=float(data.get('entrada', 0)) if data.get('entrada') else None,
            parcelas=int(data.get('parcelas')) if data.get('parcelas') else None,
            valor_parcelas=float(data.get('valor_parcelas')) if data.get('valor_parcelas') else None,
            status=data['status']
        )
        
        db.session.add(novo_registro)
        db.session.commit()
        
        # Salva no Excel
        if save_to_excel():
            return jsonify({'message': 'Registro criado e salvo no Excel com sucesso', 'registro': novo_registro.to_dict()}), 201
        else:
            return jsonify({'message': 'Registro criado no banco, mas erro ao salvar no Excel', 'registro': novo_registro.to_dict()}), 201
            
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@arthurotica_bp.route('/registros/<int:registro_id>', methods=['GET'])
def get_registro(registro_id):
    """Retorna um registro específico"""
    registro = Arthurotica.query.get_or_404(registro_id)
    return jsonify(registro.to_dict())

@arthurotica_bp.route('/registros/<int:registro_id>', methods=['PUT'])
def update_registro(registro_id):
    """Atualiza um registro"""
    try:
        registro = Arthurotica.query.get_or_404(registro_id)
        data = request.get_json()
        
        # Atualiza campos se fornecidos
        if 'id_os' in data:
            registro.id_os = data['id_os']
        if 'valor_lente' in data:
            registro.valor_lente = float(data['valor_lente'])
        if 'valor_armacao' in data:
            registro.valor_armacao = float(data['valor_armacao'])
        if 'forma_pagamento' in data:
            registro.forma_pagamento = data['forma_pagamento']
        if 'entrada' in data:
            registro.entrada = float(data['entrada']) if data['entrada'] else None
        if 'parcelas' in data:
            registro.parcelas = int(data['parcelas']) if data['parcelas'] else None
        if 'valor_parcelas' in data:
            registro.valor_parcelas = float(data['valor_parcelas']) if data['valor_parcelas'] else None
        if 'status' in data:
            registro.status = data['status']
        
        # Recalcula valor total
        registro.valor_total = registro.valor_lente + registro.valor_armacao
        
        db.session.commit()
        
        # Salva no Excel
        if save_to_excel():
            return jsonify({'message': 'Registro atualizado e salvo no Excel com sucesso', 'registro': registro.to_dict()})
        else:
            return jsonify({'message': 'Registro atualizado no banco, mas erro ao salvar no Excel', 'registro': registro.to_dict()})
            
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@arthurotica_bp.route('/registros/<int:registro_id>', methods=['DELETE'])
def delete_registro(registro_id):
    """Deleta um registro"""
    try:
        registro = Arthurotica.query.get_or_404(registro_id)
        db.session.delete(registro)
        db.session.commit()
        
        # Salva no Excel
        if save_to_excel():
            return jsonify({'message': 'Registro deletado e Excel atualizado com sucesso'})
        else:
            return jsonify({'message': 'Registro deletado do banco, mas erro ao atualizar Excel'})
            
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

