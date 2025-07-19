from flask import Blueprint, request, jsonify, send_file
from src.models.cliente import db, Cliente
from src.models.ordem_servico import OrdemServico
from src.models.receita import Receita
from src.models.armacao_lente import ArmacaoLente
from src.models.arthurotica import Arthurotica
from datetime import datetime
import openpyxl
import os
import tempfile

cliente_bp = Blueprint('cliente', __name__)

def save_clientes_to_excel():
    """Salva todos os dados de clientes para o arquivo Excel"""
    try:
        excel_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'SISTEMA-ARTHURÓTICA.xlsx')
        
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook['Cliente']
        
        # Limpa dados existentes (mantém apenas o cabeçalho)
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.value = None
        
        # Busca todos os registros do banco
        registros = Cliente.query.all()
        
        # Adiciona os dados ao Excel
        for i, registro in enumerate(registros, start=2):
            sheet.cell(row=i, column=1, value=registro.id)
            sheet.cell(row=i, column=2, value=registro.nome)
            sheet.cell(row=i, column=3, value=registro.cpf)
            sheet.cell(row=i, column=4, value=registro.data_nascimento)
            sheet.cell(row=i, column=5, value=registro.endereco)
            sheet.cell(row=i, column=6, value=registro.bairro)
            sheet.cell(row=i, column=7, value=registro.cidade)
            sheet.cell(row=i, column=8, value=registro.telefone)
        
        workbook.save(excel_path)
        return True
    except Exception as e:
        print(f"Erro ao salvar clientes no Excel: {e}")
        return False

@cliente_bp.route('/clientes', methods=['GET'])
def get_clientes():
    """Retorna todos os clientes"""
    clientes = Cliente.query.all()
    return jsonify([cliente.to_dict() for cliente in clientes])

@cliente_bp.route('/clientes', methods=['POST'])
def create_cliente():
    """Cria um novo cliente"""
    try:
        data = request.get_json()
        
        # Validação básica
        required_fields = ['nome', 'cpf']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({'error': f'Campo {field} é obrigatório'}), 400
        
        # Converte data de nascimento se fornecida
        data_nascimento = None
        if data.get('data_nascimento'):
            try:
                data_nascimento = datetime.strptime(data['data_nascimento'], '%Y-%m-%d').date()
            except ValueError:
                return jsonify({'error': 'Formato de data inválido. Use YYYY-MM-DD'}), 400
        
        # Cria novo cliente
        novo_cliente = Cliente(
            nome=data['nome'],
            cpf=data['cpf'],
            data_nascimento=data_nascimento,
            endereco=data.get('endereco'),
            bairro=data.get('bairro'),
            cidade=data.get('cidade'),
            telefone=data.get('telefone')
        )
        
        db.session.add(novo_cliente)
        db.session.commit()
        
        # Salva no Excel
        if save_clientes_to_excel():
            return jsonify({'message': 'Cliente criado e salvo no Excel com sucesso', 'cliente': novo_cliente.to_dict()}), 201
        else:
            return jsonify({'message': 'Cliente criado no banco, mas erro ao salvar no Excel', 'cliente': novo_cliente.to_dict()}), 201
            
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@cliente_bp.route('/clientes/<int:cliente_id>', methods=['GET'])
def get_cliente(cliente_id):
    """Retorna um cliente específico"""
    cliente = Cliente.query.get_or_404(cliente_id)
    return jsonify(cliente.to_dict())

@cliente_bp.route('/clientes/<int:cliente_id>', methods=['PUT'])
def update_cliente(cliente_id):
    """Atualiza um cliente"""
    try:
        cliente = Cliente.query.get_or_404(cliente_id)
        data = request.get_json()
        
        # Atualiza campos se fornecidos
        if 'nome' in data:
            cliente.nome = data['nome']
        if 'cpf' in data:
            cliente.cpf = data['cpf']
        if 'data_nascimento' in data:
            if data['data_nascimento']:
                cliente.data_nascimento = datetime.strptime(data['data_nascimento'], '%Y-%m-%d').date()
            else:
                cliente.data_nascimento = None
        if 'endereco' in data:
            cliente.endereco = data['endereco']
        if 'bairro' in data:
            cliente.bairro = data['bairro']
        if 'cidade' in data:
            cliente.cidade = data['cidade']
        if 'telefone' in data:
            cliente.telefone = data['telefone']
        
        db.session.commit()
        
        # Salva no Excel
        if save_clientes_to_excel():
            return jsonify({'message': 'Cliente atualizado e salvo no Excel com sucesso', 'cliente': cliente.to_dict()})
        else:
            return jsonify({'message': 'Cliente atualizado no banco, mas erro ao salvar no Excel', 'cliente': cliente.to_dict()})
            
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@cliente_bp.route('/clientes/<int:cliente_id>', methods=['DELETE'])
def delete_cliente(cliente_id):
    """Deleta um cliente"""
    try:
        cliente = Cliente.query.get_or_404(cliente_id)
        db.session.delete(cliente)
        db.session.commit()
        
        # Salva no Excel
        if save_clientes_to_excel():
            return jsonify({'message': 'Cliente deletado e Excel atualizado com sucesso'})
        else:
            return jsonify({'message': 'Cliente deletado do banco, mas erro ao atualizar Excel'})
            
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@cliente_bp.route('/clientes/pesquisar', methods=['GET'])
def pesquisar_clientes():
    """Pesquisa clientes por CPF ou Nome"""
    try:
        termo = request.args.get('termo', '').strip()
        
        if not termo:
            return jsonify({'error': 'Termo de pesquisa é obrigatório'}), 400
        
        # Pesquisa por CPF (busca exata) ou Nome (busca parcial)
        clientes = Cliente.query.filter(
            (Cliente.cpf == termo) | 
            (Cliente.nome.ilike(f'%{termo}%'))
        ).all()
        
        return jsonify([cliente.to_dict() for cliente in clientes])
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@cliente_bp.route('/clientes/<int:cliente_id>/relatorio', methods=['GET'])
def get_relatorio_cliente(cliente_id):
    """Retorna relatório completo do cliente com todas as informações relacionadas"""
    try:
        # Busca o cliente
        cliente = Cliente.query.get_or_404(cliente_id)
        
        # Busca todas as ordens de serviço do cliente
        ordens = OrdemServico.query.filter_by(id_cliente=cliente_id).all()
        
        relatorio = {
            'cliente': cliente.to_dict(),
            'ordens_servico': [],
            'total_compras': 0,
            'resumo_pagamentos': {
                'total_gasto': 0,
                'formas_pagamento': {},
                'status_pagamentos': {}
            }
        }
        
        for ordem in ordens:
            ordem_info = ordem.to_dict()
            
            # Busca receita da OS
            receita = Receita.query.filter_by(id_os=ordem.id).first()
            ordem_info['receita'] = receita.to_dict() if receita else None
            
            # Busca armação da OS
            armacao = ArmacaoLente.query.filter_by(id_os=ordem.id).first()
            ordem_info['armacao'] = armacao.to_dict() if armacao else None
            
            # Busca pagamento da OS
            pagamento = Arthurotica.query.filter_by(id_os=ordem.id).first()
            ordem_info['pagamento'] = pagamento.to_dict() if pagamento else None
            
            # Atualiza resumo financeiro
            if pagamento:
                relatorio['resumo_pagamentos']['total_gasto'] += pagamento.valor_total
                
                # Conta formas de pagamento
                forma = pagamento.forma_pagamento
                if forma in relatorio['resumo_pagamentos']['formas_pagamento']:
                    relatorio['resumo_pagamentos']['formas_pagamento'][forma] += 1
                else:
                    relatorio['resumo_pagamentos']['formas_pagamento'][forma] = 1
                
                # Conta status de pagamentos
                status = pagamento.status
                if status in relatorio['resumo_pagamentos']['status_pagamentos']:
                    relatorio['resumo_pagamentos']['status_pagamentos'][status] += 1
                else:
                    relatorio['resumo_pagamentos']['status_pagamentos'][status] = 1
            
            relatorio['ordens_servico'].append(ordem_info)
        
        relatorio['total_compras'] = len(ordens)
        
        return jsonify(relatorio)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@cliente_bp.route('/clientes/exportar-excel', methods=['GET'])
def exportar_clientes_excel():
    """Exporta todos os dados do sistema para uma planilha Excel"""
    try:
        # Criar um arquivo temporário
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file.close()
        
        # Criar workbook
        workbook = openpyxl.Workbook()
        
        # Remover a planilha padrão
        workbook.remove(workbook.active)
        
        # Aba Clientes
        sheet_clientes = workbook.create_sheet("Clientes")
        headers_clientes = ["ID + Nome", "CPF", "Data Nascimento", "Endereço", "Bairro", "Cidade", "Telefone"]
        sheet_clientes.append(headers_clientes)
        
        clientes = Cliente.query.all()
        for cliente in clientes:
            sheet_clientes.append([
                f'{cliente.id} - {cliente.nome}',
                cliente.cpf,
                cliente.data_nascimento.strftime('%Y-%m-%d') if cliente.data_nascimento else '',
                cliente.endereco or '',
                cliente.bairro or '',
                cliente.cidade or '',
                cliente.telefone or ''
            ])
        
        # Importar outros modelos para incluir todas as tabelas
        from src.models.ordem_servico import OrdemServico
        from src.models.receita import Receita
        from src.models.armacao_lente import ArmacaoLente
        from src.models.arthurotica import Arthurotica
        
        # Aba Ordens de Serviço
        sheet_ordens = workbook.create_sheet("Ordens de Serviço")
        headers_ordens = ["ID + Número OS", "Data Pedido", "ID Cliente"]
        sheet_ordens.append(headers_ordens)
        
        ordens = OrdemServico.query.all()
        for ordem in ordens:
            sheet_ordens.append([
                f'{ordem.id} - {ordem.numero_os}',
                ordem.data_pedido.strftime('%Y-%m-%d') if ordem.data_pedido else '',
                ordem.id_cliente
            ])
        
        # Aba Receitas
        sheet_receitas = workbook.create_sheet("Receitas")
        headers_receitas = ["ID", "ID OS", "Esférico Longe OD", "Cilíndrico Longe OD", "Eixo Longe OD", 
                           "DNP Longe OD", "Altura OD", "Adição OD", "Esférico Longe OE", "Cilíndrico Longe OE",
                           "Eixo Longe OE", "DNP Longe OE", "Altura OE", "Adição OE", "Esférico Perto OD",
                           "Cilíndrico Perto OD", "Esférico Perto OE", "Cilíndrico Perto OE"]
        sheet_receitas.append(headers_receitas)
        
        receitas = Receita.query.all()
        for receita in receitas:
            sheet_receitas.append([
                receita.id,
                receita.id_os,
                receita.esferico_longe_od,
                receita.cilindrico_longe_od,
                receita.eixo_longe_od,
                receita.dnp_longe_od,
                receita.altura_od,
                receita.adicao_od,
                receita.esferico_longe_oe,
                receita.cilindrico_longe_oe,
                receita.eixo_longe_oe,
                receita.dnp_longe_oe,
                receita.altura_oe,
                receita.adicao_oe,
                receita.esferico_perto_od,
                receita.cilindrico_perto_od,
                receita.esferico_perto_oe,
                receita.cilindrico_perto_oe
            ])
        
        # Aba Armações e Lentes
        sheet_armacoes = workbook.create_sheet("Armações e Lentes")
        headers_armacoes = ["ID", "ID OS", "Ponte", "Horizontal", "Vertical", "Diagonal Maior",
                           "Marca Armação", "Referência Armação", "Material Armação", "Lente Comprada",
                           "Tratamento", "Coloração"]
        sheet_armacoes.append(headers_armacoes)
        
        armacoes = ArmacaoLente.query.all()
        for armacao in armacoes:
            sheet_armacoes.append([
                armacao.id,
                armacao.id_os,
                armacao.ponte,
                armacao.horizontal,
                armacao.vertical,
                armacao.diagonal_maior,
                armacao.marca_armacao,
                armacao.referencia_armacao,
                armacao.material_armacao,
                armacao.lente_comprada,
                armacao.tratamento,
                armacao.coloracao
            ])
        
        # Aba Pagamentos
        sheet_pagamentos = workbook.create_sheet("Pagamentos")
        headers_pagamentos = ["ID", "ID OS", "Valor Lente", "Valor Armação", "Valor Total",
                             "Forma Pagamento", "Entrada", "Parcelas", "Valor Parcelas", "Status"]
        sheet_pagamentos.append(headers_pagamentos)
        
        pagamentos = Arthurotica.query.all()
        for pagamento in pagamentos:
            sheet_pagamentos.append([
                pagamento.id,
                pagamento.id_os,
                pagamento.valor_lente,
                pagamento.valor_armacao,
                pagamento.valor_total,
                pagamento.forma_pagamento,
                pagamento.entrada,
                pagamento.parcelas,
                pagamento.valor_parcelas,
                pagamento.status
            ])
