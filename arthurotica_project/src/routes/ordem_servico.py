from flask import Blueprint, request, jsonify
from src.models.ordem_servico import db, OrdemServico
from datetime import datetime
import openpyxl
import os

ordem_servico_bp = Blueprint('ordem_servico', __name__)

def save_ordens_to_excel():
    """Salva todos os dados de ordens de serviço para o arquivo Excel"""
    try:
        excel_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'SISTEMA-ARTHURÓTICA.xlsx')
        
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook['Ordem Serviço']
        
        # Limpa dados existentes (mantém apenas o cabeçalho)
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.value = None
        
        # Busca todos os registros do banco
        registros = OrdemServico.query.all()
        
        # Adiciona os dados ao Excel
        for i, registro in enumerate(registros, start=2):
            sheet.cell(row=i, column=1, value=registro.id)
            sheet.cell(row=i, column=2, value=registro.numero_os)
            sheet.cell(row=i, column=3, value=registro.data_pedido)
            sheet.cell(row=i, column=4, value=registro.id_cliente)
        
        workbook.save(excel_path)
        return True
    except Exception as e:
        print(f"Erro ao salvar ordens no Excel: {e}")
        return False

@ordem_servico_bp.route('/ordens', methods=['GET'])
def get_ordens():
    """Retorna todas as ordens de serviço"""
    ordens = OrdemServico.query.all()
    return jsonify([ordem.to_dict() for ordem in ordens])

@ordem_servico_bp.route('/ordens', methods=['POST'])
def create_ordem():
    """Cria uma nova ordem de serviço"""
    try:
        data = request.get_json()
        
        # Validação básica
        required_fields = ['numero_os', 'data_pedido', 'id_cliente']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({'error': f'Campo {field} é obrigatório'}), 400
        
        # Converte data do pedido
        try:
            data_pedido = datetime.strptime(data['data_pedido'], '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Formato de data inválido. Use YYYY-MM-DD'}), 400
        
        # Cria nova ordem
        nova_ordem = OrdemServico(
            numero_os=data['numero_os'],
            data_pedido=data_pedido,
            id_cliente=int(data['id_cliente'])
        )
        
        db.session.add(nova_ordem)
        db.session.commit()
        
        # Salva no Excel
        if save_ordens_to_excel():
            return jsonify({'message': 'Ordem criada e salva no Excel com sucesso', 'ordem': nova_ordem.to_dict()}), 201
        else:
            return jsonify({'message': 'Ordem criada no banco, mas erro ao salvar no Excel', 'ordem': nova_ordem.to_dict()}), 201
            
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@ordem_servico_bp.route('/ordens/<int:ordem_id>', methods=['GET'])
def get_ordem(ordem_id):
    """Retorna uma ordem específica"""
    ordem = OrdemServico.query.get_or_404(ordem_id)
    return jsonify(ordem.to_dict())

@ordem_servico_bp.route('/ordens/<int:ordem_id>', methods=['PUT'])
def update_ordem(ordem_id):
    """Atualiza uma ordem de serviço"""
    try:
        ordem = OrdemServico.query.get_or_404(ordem_id)
        data = request.get_json()
        
        # Atualiza campos se fornecidos
        if 'numero_os' in data:
            ordem.numero_os = data['numero_os']
        if 'data_pedido' in data:
            ordem.data_pedido = datetime.strptime(data['data_pedido'], '%Y-%m-%d').date()
        if 'id_cliente' in data:
            ordem.id_cliente = int(data['id_cliente'])
        
        db.session.commit()
        
        # Salva no Excel
        if save_ordens_to_excel():
            return jsonify({'message': 'Ordem atualizada e salva no Excel com sucesso', 'ordem': ordem.to_dict()})
        else:
            return jsonify({'message': 'Ordem atualizada no banco, mas erro ao salvar no Excel', 'ordem': ordem.to_dict()})
            
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@ordem_servico_bp.route('/ordens/<int:ordem_id>', methods=['DELETE'])
def delete_ordem(ordem_id):
    """Deleta uma ordem de serviço"""
    try:
        ordem = OrdemServico.query.get_or_404(ordem_id)
        db.session.delete(ordem)
        db.session.commit()
        
        # Salva no Excel
        if save_ordens_to_excel():
            return jsonify({'message': 'Ordem deletada e Excel atualizado com sucesso'})
        else:
            return jsonify({'message': 'Ordem deletada do banco, mas erro ao atualizar Excel'})
            
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

