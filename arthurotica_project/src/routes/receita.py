from flask import Blueprint, request, jsonify
from src.models.user import db
from src.models.receita import Receita
import openpyxl
import os

receita_bp = Blueprint('receita', __name__)

def save_receitas_to_excel():
    """Salva todos os dados de receitas para o arquivo Excel"""
    try:
        excel_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'SISTEMA-ARTHURÓTICA.xlsx')
        
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook['Receita']
        
        # Limpa dados existentes (mantém apenas o cabeçalho)
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.value = None
        
        # Busca todos os registros do banco
        registros = Receita.query.all()
        
        # Adiciona os dados ao Excel
        for i, registro in enumerate(registros, start=2):
            sheet.cell(row=i, column=1, value=registro.id)
            sheet.cell(row=i, column=2, value=registro.id_os)
            sheet.cell(row=i, column=3, value=registro.esferico_longe_od)
            sheet.cell(row=i, column=4, value=registro.cilindrico_longe_od)
            sheet.cell(row=i, column=5, value=registro.eixo_longe_od)
            sheet.cell(row=i, column=6, value=registro.dnp_longe_od)
            sheet.cell(row=i, column=7, value=registro.altura_od)
            sheet.cell(row=i, column=8, value=registro.adicao_od)
            sheet.cell(row=i, column=9, value=registro.esferico_longe_oe)
            sheet.cell(row=i, column=10, value=registro.cilindrico_longe_oe)
            sheet.cell(row=i, column=11, value=registro.eixo_longe_oe)
            sheet.cell(row=i, column=12, value=registro.dnp_longe_oe)
            sheet.cell(row=i, column=13, value=registro.altura_oe)
            sheet.cell(row=i, column=14, value=registro.adicao_oe)
            sheet.cell(row=i, column=15, value=registro.esferico_perto_od)
            sheet.cell(row=i, column=16, value=registro.cilindrico_perto_od)
            sheet.cell(row=i, column=17, value=registro.esferico_perto_oe)
            sheet.cell(row=i, column=18, value=registro.cilindrico_perto_oe)
        
        workbook.save(excel_path)
        return True
    except Exception as e:
        print(f"Erro ao salvar receitas no Excel: {e}")
        return False

@receita_bp.route('/receitas', methods=['GET'])
def get_receitas():
    """Retorna todas as receitas"""
    receitas = Receita.query.all()
    return jsonify([receita.to_dict() for receita in receitas])

@receita_bp.route('/receitas', methods=['POST'])
def create_receita():
    """Cria uma nova receita"""
    try:
        data = request.get_json()
        
        # Validação básica
        if 'id_os' not in data or not data['id_os']:
            return jsonify({'error': 'Campo id_os é obrigatório'}), 400
        
        # Cria nova receita
        esferico_longe_od = float(data.get("esferico_longe_od")) if data.get("esferico_longe_od") else None
        cilindrico_longe_od = float(data.get("cilindrico_longe_od")) if data.get("cilindrico_longe_od") else None
        eixo_longe_od = float(data.get("eixo_longe_od")) if data.get("eixo_longe_od") else None
        dnp_longe_od = float(data.get("dnp_longe_od")) if data.get("dnp_longe_od") else None
        altura_od = float(data.get("altura_od")) if data.get("altura_od") else None
        adicao_od = float(data.get("adicao_od")) if data.get("adicao_od") else None
        esferico_longe_oe = float(data.get("esferico_longe_oe")) if data.get("esferico_longe_oe") else None
        cilindrico_longe_oe = float(data.get("cilindrico_longe_oe")) if data.get("cilindrico_longe_oe") else None
        eixo_longe_oe = float(data.get("eixo_longe_oe")) if data.get("eixo_longe_oe") else None
        dnp_longe_oe = float(data.get("dnp_longe_oe")) if data.get("dnp_longe_oe") else None
        altura_oe = float(data.get("altura_oe")) if data.get("altura_oe") else None
        adicao_oe = float(data.get("adicao_oe")) if data.get("adicao_oe") else None

        # Aplicar regra de cilindrico não negativo
        if cilindrico_longe_od is not None: cilindrico_longe_od = abs(cilindrico_longe_od)
        if cilindrico_longe_oe is not None: cilindrico_longe_oe = abs(cilindrico_longe_oe)

        # Calcular esferico_perto e cilindrico_perto
        esferico_perto_od = (esferico_longe_od + adicao_od) if esferico_longe_od is not None and adicao_od is not None else None
        esferico_perto_oe = (esferico_longe_oe + adicao_oe) if esferico_longe_oe is not None and adicao_oe is not None else None
        cilindrico_perto_od = cilindrico_longe_od
        cilindrico_perto_oe = cilindrico_longe_oe

        nova_receita = Receita(
            id_os=int(data["id_os"]),
            esferico_longe_od=esferico_longe_od,
            cilindrico_longe_od=cilindrico_longe_od,
            eixo_longe_od=eixo_longe_od,
            dnp_longe_od=dnp_longe_od,
            altura_od=altura_od,
            adicao_od=adicao_od,
            esferico_longe_oe=esferico_longe_oe,
            cilindrico_longe_oe=cilindrico_longe_oe,
            eixo_longe_oe=eixo_longe_oe,
            dnp_longe_oe=dnp_longe_oe,
            altura_oe=altura_oe,
            adicao_oe=adicao_oe,
            esferico_perto_od=esferico_perto_od,
            cilindrico_perto_od=cilindrico_perto_od,
            esferico_perto_oe=esferico_perto_oe,
            cilindrico_perto_oe=cilindrico_perto_oe
        )
        
        db.session.add(nova_receita)
        db.session.commit()
        
        # Salva no Excel
        if save_receitas_to_excel():
            return jsonify({"message": "Receita criada e salva no Excel com sucesso", "receita": nova_receita.to_dict()}), 201
        else:
            return jsonify({"message": "Receita criada no banco, mas erro ao salvar no Excel", "receita": nova_receita.to_dict()}), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@receita_bp.route('/receitas/<int:receita_id>', methods=['GET'])
def get_receita(receita_id):
    """Retorna uma receita específica"""
    receita = Receita.query.get_or_404(receita_id)
    return jsonify(receita.to_dict())

@receita_bp.route('/receitas/<int:receita_id>', methods=['PUT'])
def update_receita(receita_id):
    """Atualiza uma receita"""
    try:
        receita = Receita.query.get_or_404(receita_id)
        data = request.get_json()
        
        # Atualiza campos se fornecidos
        fields = [
            'id_os', 'esferico_longe_od', 'cilindrico_longe_od', 'eixo_longe_od',
            'dnp_longe_od', 'altura_od', 'adicao_od', 'esferico_longe_oe',
            'cilindrico_longe_oe', 'eixo_longe_oe', 'dnp_longe_oe', 'altura_oe',
            'adicao_oe'
        ]

        for field in fields:
            if field in data:
                if field == 'id_os':
                    setattr(receita, field, int(data[field]) if data[field] else None)
                else:
                    setattr(receita, field, float(data[field]) if data[field] else None)

        # Aplicar regra de cilindrico não negativo
        if receita.cilindrico_longe_od is not None: receita.cilindrico_longe_od = abs(receita.cilindrico_longe_od)
        if receita.cilindrico_longe_oe is not None: receita.cilindrico_longe_oe = abs(receita.cilindrico_longe_oe)

        # Calcular esferico_perto e cilindrico_perto
        receita.esferico_perto_od = (receita.esferico_longe_od + receita.adicao_od) if receita.esferico_longe_od is not None and receita.adicao_od is not None else None
        receita.esferico_perto_oe = (receita.esferico_longe_oe + receita.adicao_oe) if receita.esferico_longe_oe is not None and receita.adicao_oe is not None else None
        receita.cilindrico_perto_od = receita.cilindrico_longe_od
        receita.cilindrico_perto_oe = receita.cilindrico_longe_oe
        
        db.session.commit()
        
        # Salva no Excel
        if save_receitas_to_excel():
            return jsonify({'message': 'Receita atualizada e salva no Excel com sucesso', 'receita': receita.to_dict()})
        else:
            return jsonify({'message': 'Receita atualizada no banco, mas erro ao salvar no Excel', 'receita': receita.to_dict()})
            
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

